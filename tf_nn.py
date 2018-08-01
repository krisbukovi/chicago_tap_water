# filename: tf_nn.py
# author: kris bukovi
# last modified: July 25, 2018


from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import tensorflow as tf
import numpy as np

def load_dataset():

    lat = []
    lon = []
    ave = []
    limit = []
    date = []
    first_draw = []
    fourth_draw = []
    sixth_draw = []
    five_min = []
    
    # load workbook
    wb = load_workbook('test.xlsx')

    # get first sheet name
    first_sheet = wb.sheetnames[0]

    # load worksheet
    ws = wb[first_sheet]

    for i in range(4, 849):

        try:

            # store latitude
            lat.append(float(ws['E' + str(i)].value))

            # store longitude
            lon.append(float(ws['F' + str(i)].value))

            # store average of draws
            ave.append(ws['L' + str(i)].value)

            # store over limit value
            limit.append(ws['M' + str(i)].value)

            # store date of sample
            date.append(ws['A' + str(i)].value)

            # store first draw
            first_draw.append(ws['G' + str(i)].value)

            # store fourth draw
            fourth_draw.append(ws['H' + str(i)].value)

            # store sixth draw
            sixth_draw.append(ws['I' + str(i)].value)

            # store 5 min flush
            five_min.append(ws['J' + str(i)].value)

        except Exception as e:
            print(e)

    # display dataframe 
    df = pd.DataFrame({'latitude': lat, 'longitude': lon, 'average': ave, 'limit': limit})
    # print(df)

    # display scatterplot 
    ax = df.plot.scatter(x='latitude', y='longitude', c='limit')
    #plt.scatter(df.latitude, df.longitude, c=df.limit)
    plt.show()


# create random minibatches from (X, Y)
def random_mini_batches(X, Y, mini_batch_size = 64, seed = 0):

    # number of training examples
    m = X.shape[1]

    mini_batches = []
    np.random.seed(seed)

    # Step 1: Shuffle (X, Y)
    permutation = list(np.random.permutation)
    shuffled_X = X[:, permutation]
    shuffled_Y = Y[:, permutation].reshape((Y.shape[0],m))

    # Step 2: Partition (shuffled_X, shuffled_Y) minus the end case
    # number of mini batches of size mini_batch_size in your partitioning
    num_complete_minibatches = math.floor(m/mini_batch_size)

    for k in range(0, num_complete_minibatches):
        mini_batch_X = shuffled_X[:, k * mini_batch_size : k * mini_batch_size + mini_batch_size]
        mini_batch_Y = shuffled_Y[:, k * mini_batch_size : k * mini_batch_size + mini_batch_size]
        mini_batch = (mini_batch_X, mini_batch_Y)
        mini_batches.append(mini_batch)

    return mini_batches

def convert_to_one_hot(Y, C):
    Y = np.eye(C)[Y.reshape(-1)].T
    return Y

def predict(X, parameters):
    W1 = tf.convert_to_tensor(parameters["W1"])
    b1 = tf.convert_to_tensor(parameters["b1"])
    W2 = tf.convert_to_tensor(parameters["W2"])
    b2 = tf.convert_to_tensor(parameters["b2"])
    W3 = tf.convert_to_tensor(parameters["W3"])
    b3 = tf.convert_to_tensor(parameters["b3"])

    params = {"W1": W1, "b1": b1, "W2": W2, "b2": b2, "W3": W3, "b3": b3}

    x = tf.placeholder("float", [12288, 1])

    z3 = forward_propagation_for_predict(x, params)
    p = tf.argmax(z3)

    sess = tf.Session()
    prediction = sess.run(p, feed_dict = {x: X})

    return prediction

def forward_propagation_for_predict(X, parameters):

    # retrieve the parameters from the dictionary parameters
    W1 = parameters['W1']
    b1 = parameters['b1']
    W2 = parameters['W2']
    b2 = parameters['b2']
    W3 = parameters['W3']
    b3 = parameters['b3']

    # Z1 = np.dot(W1, X) + b1
    Z1 = tf.add(tf.matmul(W1, X), b1)

    # A1 = relu(Z1)
    A1 = tf.nn.relu(Z1)

    # Z2 = np.dot(W2, X) + b2
    Z2 = tf.add(tf.matmul(W2, A1), b2)

    # A2 = relu(Z2)
    A2 = tf.nn.relu(Z2)

    # Z3 = np.dot(W3, Z2) + b3
    Z3 = tf.add(tf.matmul(W3, A2), b3)

    return Z3

def one_hot_matrix(labels, C):
    

#def ones(shape):

#def create_placeholders(n_x, n_y):

#def initialize_parameters():

#def forward_propagation(X, parameters):

#def compute_cost(Z3, Y):

#def model(X_train, Y_train, X_test, Y_test, learning_rate = 0.0001,
            #num_epochs = 1500, minibatch_size = 32, print_cost = True):

load_dataset()

