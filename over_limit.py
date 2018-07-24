# filename: over_limit.py
# author: kris bukovi
# last modified: July 24, 2018

from openpyxl import load_workbook
from geopy.geocoders import Nominatim

# create Nominatim object
geolocator = Nominatim()

# load workbook
wb = load_workbook('test.xlsx')

# get first sheet name
first_sheet = wb.sheetnames[0]

# load worksheet
ws = wb[first_sheet]

for i in range(4,849):

    num_draws = 3

    try:

        # store first draw
        first_draw = ws['G' + str(i)].value

        if first_draw == "<0.3":
            first_draw = "0.3"
        elif first_draw == "<1":
            first_draw = "1.0"
        elif first_draw == "":
            num_draws-=1

        # store fourth draw
        fourth_draw = ws['H' + str(i)].value

        if fourth_draw == "<0.3":
            fourth_draw = "0.3"
        elif fourth_draw == "<1":
            fourth_draw = "1.0"
        elif fourth_draw == "":
            num_draws-=1

        # store sixth draw
        sixth_draw = ws['I' + str(i)].value

        if sixth_draw == "<0.3":
            sixth_draw = "0.3"
        elif sixth_draw == "<1":
            sixth_draw = "1.0"
        elif sixth_draw == "":
            num_draws-=1

        # get the average 
        average = (float(first_draw) + float(fourth_draw) + float(sixth_draw)) / num_draws

        # write average to file
        ws['L' + str(i)] = average

        # write yes/no for over limit column to file
        if average > 5.0:
            ws['M' + str(i)] = "Yes"
        else:
            ws['M' + str(i)] = "No"

        # save file
        wb.save('test.xlsx')

    except Exception as e:
        print(e)
        








