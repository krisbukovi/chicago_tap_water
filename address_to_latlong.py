# filename: address_to_latlong.py
# author: kris bukovi
# last modified: July 21, 2018

from openpyxl import load_workbook
from geopy.geocoders import Nominatim

# create Nominatim object
geolocator = Nominatim()

# load workbook
wb = load_workbook('test.xlsx')

# get first sheet name
first_sheet = wb.sheetnames[0]

# load worksheet
ws = wb[first_sheet]

for i in range(4,849):

    # store address
    address = ws['B' + str(i)]

    # display address
    print(address.value)

    # create location
    location = geolocator.geocode(address.value, timeout=None)

    try:

        # get latitude
        lat = location.latitude

        # get longitude
        lon = location.longitude

        print("(" + str(lat) + "," + str(lon) + ")")

        # write latitude value to file
        ws['E' + str(i)] = lat

        # write longitude value to file
        ws['F' + str(i)] = lon

        # save file
        wb.save('test.xlsx')

    except:
        print("Can't get coordinates, skipping...")
        








